import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


COR_CABECALHO_PRINCIPAL = "0F172A"  # Slate 900 (Azul muito escuro/Chumbo)
COR_CABECALHO_TABELA    = "1E293B"  # Slate 800
COR_FUNDO_CABECALHO     = "F8FAFC"  # Fundo principal super claro
COR_TEXTO_PADRAO        = "334155"  # Cinza escuro para leitura confortável
COR_DESTAQUE            = "2563EB"  # Azul vibrante apenas para números/KPIs

COR_LINHA_PAR           = "FFFFFF"  # Branco
COR_LINHA_IMPAR         = "F1F5F9"  # Slate 100 (Cinza muito sutil)

COR_KPI_FUNDO           = "FFFFFF"
COR_KPI_BORDA           = "E2E8F0"

COR_SUCESSO_FUNDO       = "ECFDF5"  # Verde pastel
COR_SUCESSO_TEXTO       = "047857"
COR_ALERTA_FUNDO        = "FEF2F2"  # Vermelho pastel
COR_ALERTA_TEXTO        = "B91C1C"

def _fonte(tamanho=10, negrito=False, cor=COR_TEXTO_PADRAO, nome="Segoe UI") -> Font:
  
    return Font(name=nome, size=tamanho, bold=negrito, color=cor)

def _fundo(cor: str) -> PatternFill:
    return PatternFill("solid", start_color=cor, fgColor=cor)

def _alinhamento(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _borda_inferior_sutil() -> Border:
    linha = Side(style="thin", color="CBD5E1")
    return Border(bottom=linha)

def _borda_caixa() -> Border:
    linha = Side(style="thin", color=COR_KPI_BORDA)
    return Border(left=linha, right=linha, top=linha, bottom=linha)



def _preparar_aba(ws, titulo: str, subtitulo: str) -> None:
    
    ws.sheet_view.showGridLines = False

    # Fundo do cabeçalho
    for col in range(1, 15):
        for row in range(1, 4):
            ws.cell(row=row, column=col).fill = _fundo(COR_FUNDO_CABECALHO)

    # Título principal
    ws.merge_cells("B2:H2")
    ws["B2"] = titulo
    ws["B2"].font = _fonte(tamanho=18, negrito=True, cor=COR_CABECALHO_PRINCIPAL)
    ws["B2"].alignment = _alinhamento("left")
    ws.row_dimensions[2].height = 28

    # Subtítulo
    ws.merge_cells("B3:H3")
    ws["B3"] = subtitulo
    ws["B3"].font = _fonte(tamanho=10, cor="64748B")
    ws["B3"].alignment = _alinhamento("left")
    ws.row_dimensions[3].height = 18

def _ajustar_colunas(ws, larguras: list[int]) -> None:
   
    ws.column_dimensions['A'].width = 3 
    for i, largura in enumerate(larguras, start=2):
        ws.column_dimensions[get_column_letter(i)].width = largura

def _desenhar_cabecalho_tabela(ws, linha: int, colunas: list[str]) -> None:
    for col_idx, texto in enumerate(colunas, start=2):
        cel = ws.cell(row=linha, column=col_idx, value=texto)
        cel.font = _fonte(tamanho=10, negrito=True, cor="FFFFFF")
        cel.fill = _fundo(COR_CABECALHO_TABELA)
        cel.alignment = _alinhamento("left")
        ws.row_dimensions[linha].height = 20

def _desenhar_linha_dados(ws, linha: int, valores: list, alternada: bool = False, alerta: bool = False) -> None:
    cor_fundo = COR_LINHA_IMPAR if alternada else COR_LINHA_PAR
    if alerta:
        cor_fundo = COR_ALERTA_FUNDO
        
    for col_idx, valor in enumerate(valores, start=2): 
        cel = ws.cell(row=linha, column=col_idx, value=valor)
        cel.font = _fonte(tamanho=10, cor=COR_ALERTA_TEXTO if alerta else COR_TEXTO_PADRAO)
        cel.fill = _fundo(cor_fundo)
        cel.alignment = _alinhamento("left")
        cel.border = _borda_inferior_sutil()
    ws.row_dimensions[linha].height = 18


def _aba_resumo(wb: Workbook, metricas: dict) -> None:
    ws = wb.active
    ws.title = "Resumo Executivo"
    _preparar_aba(ws, "Resumo Executivo", "Análise de performance e saúde das transações.")

    kpis = [
        ("Faturamento Total",    f"R$ {metricas['faturamento_total']:,.2f}",    COR_DESTAQUE),
        ("Ticket Médio",         f"R$ {metricas['ticket_medio']:,.2f}",         COR_TEXTO_PADRAO),
        ("Taxa de Aprovação",    f"{metricas['taxa_aprovacao']}%",              COR_SUCESSO_TEXTO),
        ("Total de Transações",  str(metricas['total_transacoes']),             COR_TEXTO_PADRAO),
        ("Aprovadas",            str(metricas['total_aprovadas']),              COR_SUCESSO_TEXTO),
        ("Recusadas (Luhn)",     str(metricas['total_recusadas']),              COR_ALERTA_TEXTO),
    ]

    linha_atual = 5

    for i, (rotulo, valor, cor_valor) in enumerate(kpis):
        col_inicio = (i % 3) * 2 + 2  
        col_fim    = col_inicio + 1

        if i > 0 and i % 3 == 0:
            linha_atual += 4

   
        ws.merge_cells(start_row=linha_atual, start_column=col_inicio, end_row=linha_atual, end_column=col_fim)
        cel_rotulo = ws.cell(row=linha_atual, column=col_inicio, value=rotulo.upper())
        cel_rotulo.font = _fonte(tamanho=8, negrito=True, cor="94A3B8")
        cel_rotulo.fill = _fundo(COR_KPI_FUNDO)
        cel_rotulo.alignment = _alinhamento("center")
        cel_rotulo.border = Border(left=Side(style="thin", color=COR_KPI_BORDA), top=Side(style="thin", color=COR_KPI_BORDA), right=Side(style="thin", color=COR_KPI_BORDA))
        ws.row_dimensions[linha_atual].height = 20


        ws.merge_cells(start_row=linha_atual + 1, start_column=col_inicio, end_row=linha_atual + 1, end_column=col_fim)
        cel_valor = ws.cell(row=linha_atual + 1, column=col_inicio, value=valor)
        cel_valor.font = _fonte(tamanho=16, negrito=True, cor=cor_valor)
        cel_valor.fill = _fundo(COR_KPI_FUNDO)
        cel_valor.alignment = _alinhamento("center")
        cel_valor.border = Border(left=Side(style="thin", color=COR_KPI_BORDA), bottom=Side(style="thin", color=COR_KPI_BORDA), right=Side(style="thin", color=COR_KPI_BORDA))
        ws.row_dimensions[linha_atual + 1].height = 30

    _ajustar_colunas(ws, [18, 5, 18, 5, 18, 5])

def _aba_periodo(wb: Workbook, metricas: dict) -> None:
    ws = wb.create_sheet("Por Período")
    df = metricas["faturamento_por_periodo"]
    _preparar_aba(ws, "Faturamento por Período", "Receita mensal consolidada (apenas transações aprovadas).")

    _desenhar_cabecalho_tabela(ws, 5, ["Período", "Faturamento"])

    for i, row in df.iterrows():
        linha = 6 + list(df.index).index(i)
        _desenhar_linha_dados(ws, linha, [str(row["periodo"]), row["faturamento"]], alternada=linha % 2 == 0)
        ws.cell(row=linha, column=3).number_format = 'R$ #,##0.00'

    chart = BarChart()
    chart.type = "col"
    chart.style = 2  
    chart.title = None 
    chart.y_axis.title = "Faturamento (R$)"
    chart.width = 18
    chart.height = 10
    

    dados = Reference(ws, min_col=3, min_row=5, max_row=5 + len(df))
    categorias = Reference(ws, min_col=2, min_row=6, max_row=5 + len(df))
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(categorias)
    
  
    chart.legend = None 

    ws.add_chart(chart, "E5")
    _ajustar_colunas(ws, [20, 20])

def _aba_produtos(wb: Workbook, metricas: dict) -> None:
    ws = wb.create_sheet("Top Produtos")
    df = metricas["top_produtos"]
    _preparar_aba(ws, "Performance de Produtos", "Curva ABC de produtos por volume de faturamento.")

    _desenhar_cabecalho_tabela(ws, 5, ["Ranking", "Produto", "Faturamento", "Volume (Qtd)"])

    for i, row in enumerate(df.itertuples(), start=1):
        linha = 5 + i
        _desenhar_linha_dados(
            ws, linha,
            [f"#{i}", row.produto, row.faturamento, int(row.quantidade_vendida)],
            alternada=i % 2 == 0
        )
        ws.cell(row=linha, column=4).number_format = 'R$ #,##0.00'
      
        ws.cell(row=linha, column=2).alignment = _alinhamento("center")
        ws.cell(row=linha, column=5).alignment = _alinhamento("center")

    _ajustar_colunas(ws, [10, 35, 20, 15])

def _aba_auditoria(wb: Workbook, df_suspeitos: pd.DataFrame) -> None:
    ws = wb.create_sheet("Auditoria")
    _preparar_aba(ws, "Relatório de Auditoria", "Transações bloqueadas pelo Algoritmo de Luhn (Potencial Fraude).")

    if df_suspeitos.empty:
        ws["B5"] = "Nenhuma transação suspeita encontrada na base de dados."
        ws["B5"].font = _fonte(tamanho=11, negrito=True, cor=COR_SUCESSO_TEXTO)
        return

    colunas = ["ID", "Data", "Nº Cartão", "Valor", "Status", "Motivo"]
    _desenhar_cabecalho_tabela(ws, 5, colunas)

    for i, row in enumerate(df_suspeitos.itertuples(), start=1):
        linha = 5 + i
        data_str = str(row.data)[:10] if hasattr(row, "data") else ""
        valores = [
            getattr(row, "id_transacao", ""),
            data_str,
            getattr(row, "numero_cartao", ""),
            getattr(row, "valor", ""),
            getattr(row, "status", ""),
            getattr(row, "motivo_auditoria", "Falha de validação (Luhn)"),
        ]
       
        _desenhar_linha_dados(ws, linha, valores, alerta=True)
        ws.cell(row=linha, column=5).number_format = 'R$ #,##0.00'

    _ajustar_colunas(ws, [12, 14, 20, 15, 12, 35])

def _aba_verificadas(wb: Workbook, df_validos: pd.DataFrame) -> None:
    from brand_detector import enriquecer_dataframe
    ws = wb.create_sheet("Transações Verificadas")
    df = enriquecer_dataframe(df_validos)
    _preparar_aba(ws, "Base Verificada", "Lista completa de transações aprovadas e mascaradas para segurança (LGPD).")

    colunas = ["ID", "Data", "Bandeira", "Cartão (Mascarado)", "Valor", "Método", "Status"]
    _desenhar_cabecalho_tabela(ws, 5, colunas)

    for i, row in enumerate(df.itertuples(), start=1):
        linha = 5 + i
        data_str = str(row.data)[:10] if hasattr(row, "data") else ""
        bandeira_txt = f"{getattr(row, 'bandeira_emoji', '')} {getattr(row, 'bandeira', '')}"
        valores = [
            getattr(row, "id_transacao", ""),
            data_str,
            bandeira_txt,
            getattr(row, "cartao_mascarado", ""),
            getattr(row, "valor", ""),
            getattr(row, "tipo_pagamento", ""),
            getattr(row, "status", ""),
        ]
        _desenhar_linha_dados(ws, linha, valores, alternada=i % 2 == 0)
        ws.cell(row=linha, column=6).number_format = 'R$ #,##0.00'

    _ajustar_colunas(ws, [12, 14, 18, 22, 15, 15, 12])

def gerar_relatorio(metricas: dict, df_suspeitos: pd.DataFrame, df_validos: pd.DataFrame = None) -> bytes:
    wb = Workbook()
    
    _aba_resumo(wb, metricas)
    _aba_periodo(wb, metricas)
    _aba_produtos(wb, metricas)
    _aba_auditoria(wb, df_suspeitos)
    if df_validos is not None:
        _aba_verificadas(wb, df_validos)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()